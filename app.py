import time
import streamlit as st
import pandas as pd
import tempfile
import os
import traceback
from openpyxl import load_workbook
from yahoo import get_yahoo_finance_data
from ism_script import update_excel_data as update_manufacturing
from ism_service_script import update_excel_data as update_service

# Set page config FIRST
st.set_page_config(
    page_title="Fundamental Report Generator",
    layout="wide",
    page_icon="📈",
    initial_sidebar_state="expanded"
)

# Custom CSS with professional styling
st.markdown("""
<style>
    /* Main background */
    .stApp {
        background: #f8f9fa;
    }
    
    /* Card-style containers */
    .custom-container {
        background: white;
        border-radius: 15px;
        padding: 25px;
        margin: 15px 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border: 1px solid #e9ecef;
    }
    
    /* Header styling */
    .header-style {
        text-align: center;
        color: #2c3e50;
        margin-bottom: 30px;
        font-weight: 600;
    }
    
    /* Step-by-step guide */
    .step-guide {
        border-left: 4px solid #3498db;
        padding-left: 20px;
        margin: 20px 0;
    }
    
    /* Progress bar customization */
    .stProgress > div > div > div {
        background-color: #fffff !important;
    }
    
    /* Button styling */
    .stButton>button {
        background-color: #3498db !important;
        transition: all 0.3s ease;
    }
    
    .stButton>button:hover {
        background-color: #2980b9 !important;
        transform: scale(1.02);
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'header_row' not in st.session_state:
    st.session_state.header_row = None
if 'ticker_col' not in st.session_state:
    st.session_state.ticker_col = None
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'download_data' not in st.session_state:
    st.session_state.download_data = None

def find_ticker_position(uploaded_file):
    """Find Ticker column position using openpyxl"""
    try:
        wb = load_workbook(uploaded_file, read_only=True)
        ws = wb.active
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for col_idx, value in enumerate(row, 1):
                if value and str(value).strip().lower() == "ticker":
                    return row_idx, col_idx
        return None, None
    except Exception as e:
        st.error(f"Error scanning file: {str(e)}")
        return None, None

def validate_file(uploaded_file, report_type):
    """Validate file based on report type"""
    try:
        if report_type == "Yahoo Data":
            header_row, ticker_col = find_ticker_position(uploaded_file)
            if not header_row or not ticker_col:
                return False, "Ticker column not found in file"
            st.session_state.header_row = header_row
            st.session_state.ticker_col = ticker_col
            return True, ""
        else:
            df = pd.read_excel(uploaded_file)
            required_col = "ISM Manufacturing" if report_type == "Manufacturing" else "ISM Service"
            if required_col not in df.columns:
                return False, f"First column must be '{required_col}'"
            return True, ""
    except Exception as e:
        return False, f"File error: {str(e)}"

def process_yahoo_data(uploaded_file):
    """Process and enhance Excel file with Yahoo data"""
    try:
        wb = load_workbook(uploaded_file)
        ws = wb.active
        header_row = st.session_state.header_row
        ticker_col = st.session_state.ticker_col

        # Collect ticker rows
        tickers = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            ticker_cell = ws.cell(row=row_idx, column=ticker_col)
            if ticker_cell.value and not str(ticker_cell.value).startswith("->"):
                tickers.append((row_idx, ticker_cell.value.strip()))

        if not tickers:
            raise ValueError("No valid tickers found")

        # Define new columns to add
        new_columns = [
            'Last Price', 'Market Cap', 'Shares Outstanding', 'Float',
            'Sector', 'Industry', '50D MA', '% vs 50D MA', '200D MA', 
            '% vs 200D MA', 'Avg Volume 10D', 'Avg Volume 3M', 
            'Shares Short', 'Short Ratio', 'Short % Float', 'Days to Cover',
            'Next Earnings Date', 'Recommendations Total', 'Strong Sell',
            'Sell', 'Hold', 'Buy', 'Strong Buy', '% of Strong Sell',
            '% of Sell', '% Hold', '% Buy', '% Strong Buy', 'Recommendation %',
            'Current Price', 'Low Target', 'Avg Target', 'High Target',
            'Low Below Current Abs %', 'High Above Current Abs %', 'Last Updated'
        ]

        # Add headers
        start_col = ws.max_column + 1
        for idx, col in enumerate(new_columns, start=start_col):
            ws.cell(row=header_row, column=idx, value=col)

        # Process each ticker
        progress_bar = st.progress(0)
        total_tickers = len(tickers)
        
        for i, (row_idx, ticker) in enumerate(tickers):
            data = get_yahoo_finance_data(ticker)
            if data:
                for col_idx, key in enumerate(new_columns, start=start_col):
                    ws.cell(row=row_idx, column=col_idx, value=data.get(key, ""))
            time.sleep(1)
            progress_bar.progress((i+1)/total_tickers)

        # Save processed file
        temp_path = os.path.join(tempfile.gettempdir(), "enhanced_data.xlsx")
        wb.save(temp_path)
        return temp_path

    except Exception as e:
        st.error(f"Processing error: {str(e)}")
        raise

def main():
    # Main header
    st.markdown("""
    <div class="header-style">
        <h1>📊 Fundamental Report Generator</h1>
        <p>Intelligent Financial Data Processing Platform</p>
    </div>
    """, unsafe_allow_html=True)

    # Tutorial section
    with st.expander("📘 How It Works - Click to Expand", expanded=True):
        cols = st.columns(3)
        with cols[0]:
            st.markdown("""
            <div class="step-guide">
                <h4>Step 1: Upload</h4>
                <p>Select report type and upload your Excel template</p>
            </div>
            """, unsafe_allow_html=True)
        with cols[1]:
            st.markdown("""
            <div class="step-guide">
                <h4>Step 2: Process</h4>
                <p>System automatically enhances your data</p>
            </div>
            """, unsafe_allow_html=True)
        with cols[2]:
            st.markdown("""
            <div class="step-guide">
                <h4>Step 3: Download</h4>
                <p>Get your updated file with fresh insights</p>
            </div>
            """, unsafe_allow_html=True)

    # Main functionality container
    with st.container():
        st.markdown('<div class="custom-container">', unsafe_allow_html=True)
        
        # Report type selection
        report_type = st.radio(
            "Select Report Type:", 
            ["Manufacturing", "Services", "Yahoo Data"],
            index=2,
            horizontal=True
        )

        # File upload section
        uploaded_file = st.file_uploader(
            f"📤 Upload {report_type} Report",
            type=["xlsx"],
            help="Drag and drop your Excel file here"
        )

        # Processing logic
        if report_type == "Yahoo Data":
            if st.button("🚀 Process Financial Data"):
                if not uploaded_file:
                    st.error("Please upload a file first")
                else:
                    try:
                        is_valid, msg = validate_file(uploaded_file, report_type)
                        if not is_valid:
                            st.error(msg)
                            return
                        
                        with st.spinner("Processing financial data..."):
                            processed_path = process_yahoo_data(uploaded_file)
                            
                            with open(processed_path, "rb") as f:
                                st.session_state.download_data = f.read()
                            
                            st.session_state.processed = True
                            st.success("Data processing completed successfully!")

                    except Exception as e:
                        st.error(f"Error: {str(e)}")
                        st.error(f"Details: {traceback.format_exc()}")
                    finally:
                        try:
                            if processed_path and os.path.exists(processed_path):
                                os.remove(processed_path)
                        except:
                            pass

        else:  # Manufacturing/Services
            default_url = (
                "https://www.ismworld.org/supply-management-news-and-reports/reports/ism-report-on-business/pmi/march/"
                if report_type == "Manufacturing" else
                "https://www.ismworld.org/supply-management-news-and-reports/reports/ism-report-on-business/services/march/"
            )
            
            url = st.text_input("🌐 ISM Report URL", value=default_url)
            
            if st.button("🔄 Update ISM Report"):
                if not uploaded_file:
                    st.error("Please upload a file first")
                else:
                    try:
                        is_valid, msg = validate_file(uploaded_file, report_type)
                        if not is_valid:
                            st.error(msg)
                            return

                        with st.spinner("Updating ISM report..."):
                            temp_path = os.path.join(tempfile.gettempdir(), f"temp_{report_type}.xlsx")
                            
                            with open(temp_path, "wb") as f:
                                f.write(uploaded_file.getvalue())
                            
                            if report_type == "Manufacturing":
                                update_manufacturing(temp_path, url)
                            else:
                                update_service(temp_path, url)
                            
                            with open(temp_path, "rb") as f:
                                st.session_state.download_data = f.read()
                            
                            st.session_state.processed = True
                            st.success("Report updated successfully!")

                    except Exception as e:
                        st.error(f"Error: {str(e)}")
                        st.error(f"Details: {traceback.format_exc()}")
                    finally:
                        try:
                            if os.path.exists(temp_path):
                                os.remove(temp_path)
                        except:
                            pass

        st.markdown('</div>', unsafe_allow_html=True)

    # Download section
    if st.session_state.processed:
        with st.container():
            st.markdown('<div class="custom-container">', unsafe_allow_html=True)
            
            download_filename = {
                "Manufacturing": "Enhanced_Manufacturing_Report.xlsx",
                "Services": "Enhanced_Service_Report.xlsx",
                "Yahoo Data": "Financial_Insights_Report.xlsx"
            }[report_type]
            
            st.download_button(
                label="📥 Download Enhanced Report",
                data=st.session_state.download_data,
                file_name=download_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.markdown('</div>', unsafe_allow_html=True)

    # Footer
    st.markdown("""
    <div style="text-align: center; margin-top: 50px; color: #666; font-size: 0.9em;">
        <hr>
        <p> 2024 OmniDataX Analytics • 📧 admin@omnidatax.com • 🔒 Secure Processing</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()