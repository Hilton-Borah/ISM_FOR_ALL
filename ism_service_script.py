import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

STANDARD_INDUSTRIES = [
    "Accommodation & Food Services",
    "Agriculture, Forestry, Fishing & Hunting",
    "Arts, Entertainment & Recreation",
    "Construction",
    "Educational Services",
    "Finance & Insurance",
    "Health Care & Social Assistance",
    "Information",
    "Management of Companies & Support Services",
    "Mining",
    "Other Services",
    "Professional, Scientific & Technical Services",
    "Public Administration",
    "Real Estate, Rental & Leasing",
    "Retail Trade",
    "Transportation & Warehousing",
    "Utilities",
    "Wholesale Trade"
]

def extract_report_month(soup):
    try:
        full_text = soup.get_text(separator=' ', strip=True)
        match = re.search(
            r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', 
            full_text
        )
        if match:
            month = match.group(1)[:3]
            year = match.group(2)[-2:]
            return f"{month}-{year}".title()
        return None
    except Exception as e:
        print(f"Month extraction error: {str(e)}")
        return None

def extract_industry_data(soup):
    try:
        text = ' '.join(soup.get_text(separator=' ', strip=True).split())
        
        # Improved growth pattern
        growth_match = re.search(
            r'industries reporting growth.*?:\s*(.*?)(?:\.|industries reporting contraction|$)',
            text, re.IGNORECASE
        )
        growth = [i.strip().replace('and ', '') for i in re.split(r'[;,]', growth_match.group(1))] if growth_match else []
        
        # Fixed contraction pattern - added space and made more robust
        contraction_match = re.search(
            r'(?:industries|sectors) reporting (?:a )?contraction.*?:\s*(.*?)(?:\.|industries reporting growth|$)',
            text, re.IGNORECASE
        )
        contraction = [i.strip().replace('and ', '') for i in re.split(r'[;,]', contraction_match.group(1))] if contraction_match else []
        
        # Clean up any empty strings
        growth = [g for g in growth if g]
        contraction = [c for c in contraction if c]
        
        return growth, contraction
    except Exception as e:
        print(f"Industry extraction error: {str(e)}")
        return None, None

def match_industry_name(report_name, standard_names):
    report_name = report_name.strip().lower()
    for std_name in standard_names:
        if report_name == std_name.lower():
            return std_name
    for std_name in standard_names:
        std_lower = std_name.lower()
        if report_name in std_lower or std_lower in report_name:
            return std_name
    clean_report = re.sub(r'[^a-z0-9\s]', '', report_name)
    for std_name in standard_names:
        clean_std = re.sub(r'[^a-z0-9\s]', '', std_name.lower())
        if clean_report == clean_std:
            return std_name
    return None

def apply_historical_formatting(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Find all month columns
    month_data = {}
    for col in ws.iter_cols(1, ws.max_column):
        header = col[0].value
        if not header:
            continue
            
        if " Rank" in header:
            month = header.replace(" Rank", "")
            month_data.setdefault(month, {})["rank_col"] = col[0].column
        elif " Status" in header:
            month = header.replace(" Status", "")
            month_data.setdefault(month, {})["status_col"] = col[0].column

    # Process each month's columns
    for month, cols in month_data.items():
        if 'rank_col' not in cols or 'status_col' not in cols:
            continue

        rank_col = cols['rank_col']
        status_col = cols['status_col']
        
        # Get rank values and calculate min/max
        rank_values = []
        for cell in ws[get_column_letter(rank_col)][1:]:  # Skip header
            try:
                value = float(re.search(r'-?\d+', str(cell.value)).group())
                rank_values.append(value)
            except:
                rank_values.append(0)
        
        if not rank_values:
            continue
            
        max_rank = max(rank_values) if rank_values else 1
        min_rank = min(rank_values) if rank_values else -1
        max_abs = max(abs(max_rank), abs(min_rank)) or 1

        # Formatting parameters
        color_intensity = 150
        base_brightness = 205

        for idx, (rank_cell, status_cell) in enumerate(zip(
            ws[get_column_letter(rank_col)][1:],
            ws[get_column_letter(status_col)][1:]
        )):
            try:
                raw_value = re.search(r'-?\d+', str(rank_cell.value)).group()
                rank_value = float(raw_value)
            except:
                continue

            # Rank column font coloring
            if rank_value > 0:
                rank_cell.font = Font(color='006100')
            elif rank_value < 0:
                rank_cell.font = Font(color='9C0006')
            else:
                rank_cell.font = Font(color='FFC000')

            # Status column gradient coloring
            status_text = status_cell.value
            if "Growth" in status_text:
                intensity = abs(rank_value / max_rank)
                green = int(base_brightness - (color_intensity * intensity))
                fill = PatternFill(
                    start_color=f"FF{green:02X}FF{green:02X}",
                    fill_type="solid"
                )
            elif "Contraction" in status_text:
                intensity = abs(rank_value / min_rank)
                red = int(base_brightness - (color_intensity * intensity))
                fill = PatternFill(
                    start_color=f"FFFF{red:02X}{red:02X}",
                    fill_type="solid"
                )
            else:
                fill = PatternFill(
                    start_color="FFEB9C",
                    fill_type="solid"
                )


            status_cell.fill = fill

    wb.save(file_path)
    wb.close()
    print("✅ Formatting applied to all historical columns")

def update_excel_data(file_path, url):
    # Preserve existing data
    if os.path.exists(file_path):
        df = pd.read_excel(file_path, dtype=str).dropna(how='all')
        existing_industries = df['ISM Service'].str.strip().tolist()
    else:
        df = pd.DataFrame({"ISM Service": STANDARD_INDUSTRIES})
        existing_industries = STANDARD_INDUSTRIES.copy()

    headers = {'User-Agent': 'Mozilla/5.0'}
    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        # Extract report data
        report_month = extract_report_month(soup)
        if not report_month:
            print("❌ Could not extract report month")
            return df

        growth, contraction = extract_industry_data(soup)
        if not growth and not contraction:
            print("❌ Could not extract industry data")
            return df

        # Add new industries
        new_industries = [
            industry for industry in growth + contraction
            if not match_industry_name(industry, existing_industries)
        ]
        if new_industries:
            new_rows = pd.DataFrame({
                'ISM Service': new_industries,
                **{col: '' for col in df.columns if col != 'ISM Service'}
            })
            df = pd.concat([df, new_rows], ignore_index=True)

        # Remove existing month data if present
        df = df.drop(columns=[f"{report_month} Rank", f"{report_month} Status"], errors='ignore')

        # Create a mapping of industry to its position in growth/contraction lists
        growth_positions = {}
        for idx, industry in enumerate(growth, 1):
            matched = match_industry_name(industry, df['ISM Service'].tolist())
            if matched and matched not in growth_positions:
                growth_positions[matched] = idx

        contraction_positions = {}
        for idx, industry in enumerate(contraction, 1):
            matched = match_industry_name(industry, df['ISM Service'].tolist())
            if matched and matched not in contraction_positions:
                contraction_positions[matched] = idx

        # Create rankings and statuses with correct consecutive numbering
        rankings, statuses = [], []
        
        # Get sorted growth industries by their mention order
        sorted_growth = sorted(growth_positions.items(), key=lambda x: x[1])
        # Reverse the ranking order for growth industries (highest number for first mentioned)
        growth_count = len(sorted_growth)
        growth_ranks = {industry: growth_count - rank + 1 for rank, (industry, _) in enumerate(sorted_growth, 1)}
        
        # # Get sorted contraction industries by their mention order (keep original order)
        # sorted_contraction = sorted(contraction_positions.items(), key=lambda x: x[1])
        # contraction_ranks = {industry: rank for rank, (industry, _) in enumerate(sorted_contraction, 1)}

        # Get sorted contraction industries by their mention order (now also reversed)
        sorted_contraction = sorted(contraction_positions.items(), key=lambda x: x[1])
        contraction_count = len(sorted_contraction)
        contraction_ranks = {industry: contraction_count - rank + 1 for rank, (industry, _) in enumerate(sorted_contraction, 1)}

        for industry in df['ISM Service']:
            if industry in growth_ranks:
                rank = growth_ranks[industry]
                rankings.append(f"{rank} ↑")
                statuses.append("Growth")
            elif industry in contraction_ranks:
                rank = contraction_ranks[industry]
                rankings.append(f"-{rank} ↓")
                statuses.append("Contraction")
            else:
                rankings.append("0 →")
                statuses.append("Neutral")

        # Update dataframe
        df[f"{report_month} Rank"] = rankings
        df[f"{report_month} Status"] = statuses

        # Save and format
        df.to_excel(file_path, index=False)
        apply_historical_formatting(file_path)
        
        # Close workbook explicitly
        wb = load_workbook(file_path)
        wb.close()
        
        return df

    except Exception as e:
        print(f"Error: {str(e)}")
        return df
    
if __name__ == "__main__":
    # For standalone use with fixed file
    EXCEL_PATH = r"ISM Service.xlsx"
    URL = "https://www.ismworld.org/supply-management-news-and-reports/reports/ism-report-on-business/services/march/"
    print("🚀 Starting ISM Report Update...")
    updated_df = update_excel_data(EXCEL_PATH, URL)
    print("\n✅ Process completed.")