import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

STANDARD_INDUSTRIES = [
    "Electrical Equipment, Appliances & Components",
    "Textile Mills", "Furniture & Related Products", "Machinery",
    "Fabricated Metal Products", "Primary Metals", "Miscellaneous Manufacturing",
    "Chemical Products", "Plastics & Rubber Products", "Food, Beverage & Tobacco Products",
    "Computer & Electronic Products", "Nonmetallic Mineral Products",
    "Apparel, Leather & Allied Products", "Transportation Equipment",
    "Paper Products", "Petroleum & Coal Products", "Printing & Related Support Activities",
    "Wood Products"
]

def extract_report_month(soup):
    try:
        full_text = soup.get_text(separator=' ', strip=True)
        match = re.search(
            r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', 
            full_text
        )
        if match:
            month = match.group(1)[:3]
            year = match.group(2)[-2:]
            return f"{month}-{year}".title()
        return None
    except Exception as e:
        print(f"Month extraction error: {str(e)}")
        return None

def extract_industry_data(soup):
    try:
        text = ' '.join(soup.get_text(separator=' ', strip=True).split())
        
        growth_match = re.search(
            r'industries reporting growth.*?:\s*(.*?)(?:\.|industries reporting contraction|$)',
            text, re.IGNORECASE
        )
        growth = [i.strip() for i in re.split(r'[;]', growth_match.group(1))] if growth_match else []
        
        contraction_match = re.search(
            r'industries reporting contraction.*?:\s*(.*?)(?:\.|$)',
            text, re.IGNORECASE
        )
        contraction = [i.strip() for i in re.split(r'[;]', contraction_match.group(1))] if contraction_match else []
        
        return growth, contraction
    except Exception as e:
        print(f"Industry extraction error: {str(e)}")
        return None, None

def match_industry_name(report_name, standard_names):
    report_name = report_name.strip().lower()
    for std_name in standard_names:
        if report_name == std_name.lower():
            return std_name
    for std_name in standard_names:
        std_lower = std_name.lower()
        if report_name in std_lower or std_lower in report_name:
            return std_name
    clean_report = re.sub(r'[^a-z0-9\s]', '', report_name)
    for std_name in standard_names:
        clean_std = re.sub(r'[^a-z0-9\s]', '', std_name.lower())
        if clean_report == clean_std:
            return std_name
    return None

def apply_historical_formatting(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Find all month columns
    month_data = {}
    for col in ws.iter_cols(1, ws.max_column):
        header = col[0].value
        if not header:
            continue
            
        if " Rank" in header:
            month = header.replace(" Rank", "")
            month_data.setdefault(month, {})["rank_col"] = col[0].column
        elif " Status" in header:
            month = header.replace(" Status", "")
            month_data.setdefault(month, {})["status_col"] = col[0].column

    # Process each month's columns
    for month, cols in month_data.items():
        if 'rank_col' not in cols or 'status_col' not in cols:
            continue

        rank_col = cols['rank_col']
        status_col = cols['status_col']
        
        # Get rank values and calculate min/max
        rank_values = []
        for cell in ws[get_column_letter(rank_col)][1:]:  # Skip header
            try:
                value = float(re.search(r'-?\d+', str(cell.value)).group())
                rank_values.append(value)
            except:
                rank_values.append(0)
        
        if not rank_values:
            continue
            
        max_rank = max(rank_values) if rank_values else 1
        min_rank = min(rank_values) if rank_values else -1
        max_abs = max(abs(max_rank), abs(min_rank)) or 1

        # Formatting parameters
        color_intensity = 150
        base_brightness = 205

        for idx, (rank_cell, status_cell) in enumerate(zip(
            ws[get_column_letter(rank_col)][1:],
            ws[get_column_letter(status_col)][1:]
        )):
            try:
                raw_value = re.search(r'-?\d+', str(rank_cell.value)).group()
                rank_value = float(raw_value)
            except:
                continue

            # Rank column font coloring
            if rank_value > 0:
                rank_cell.font = Font(color='006100')
            elif rank_value < 0:
                rank_cell.font = Font(color='9C0006')
            else:
                rank_cell.font = Font(color='FFC000')

            # Status column gradient coloring
            status_text = status_cell.value
            if "Growth" in status_text:
                intensity = abs(rank_value / max_rank)
                green = int(base_brightness - (color_intensity * intensity))
                fill = PatternFill(
                    start_color=f"FF{green:02X}FF{green:02X}",
                    fill_type="solid"
                )
            elif "Contraction" in status_text:
                intensity = abs(rank_value / min_rank)
                red = int(base_brightness - (color_intensity * intensity))
                fill = PatternFill(
                    start_color=f"FFFF{red:02X}{red:02X}",
                    fill_type="solid"
                )
            else:
                fill = PatternFill(
                    start_color="FFEB9C",
                    fill_type="solid"
                )
            
            status_cell.fill = fill

    wb.save(file_path)
    wb.close()
    print("✅ Formatting applied to all historical columns")

def update_excel_data(file_path, url):
    # Preserve existing data
    if os.path.exists(file_path):
        df = pd.read_excel(file_path, dtype=str).dropna(how='all')
        existing_industries = df['ISM Manufacturing'].str.strip().tolist()
    else:
        df = pd.DataFrame({"ISM Manufacturing": STANDARD_INDUSTRIES})
        existing_industries = STANDARD_INDUSTRIES.copy()

    headers = {'User-Agent': 'Mozilla/5.0'}
    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        # Extract report data
        report_month = extract_report_month(soup)
        if not report_month:
            print("❌ Could not extract report month")
            return df

        growth, contraction = extract_industry_data(soup)
        if not growth and not contraction:
            print("❌ Could not extract industry data")
            return df

        # Add new industries
        new_industries = [
            industry for industry in growth + contraction
            if not match_industry_name(industry, existing_industries)
        ]
        if new_industries:
            new_rows = pd.DataFrame({
                'ISM Manufacturing': new_industries,
                **{col: '' for col in df.columns if col != 'ISM Manufacturing'}
            })
            df = pd.concat([df, new_rows], ignore_index=True)

        # Remove existing month data if present
        df = df.drop(columns=[f"{report_month} Rank", f"{report_month} Status"], errors='ignore')

        # Create rankings and statuses
        rankings, statuses = [], []
        for industry in df['ISM Manufacturing']:
            matched_growth = next(
                (i for i, g in enumerate(reversed(growth), 1) 
                if match_industry_name(g, [industry])),
                None
            )
            matched_contraction = next(
                (i for i, c in enumerate(reversed(contraction), 1) 
                if not matched_growth and match_industry_name(c, [industry])),
                None
            )

            if matched_growth:
                rankings.append(f"{matched_growth} ↑")
                statuses.append("Growth")
            elif matched_contraction:
                rankings.append(f"-{matched_contraction} ↓")
                statuses.append("Contraction")
            else:
                rankings.append("0 →")
                statuses.append("Neutral")

        # Update dataframe
        df[f"{report_month} Rank"] = rankings
        df[f"{report_month} Status"] = statuses

        # Save and format
        df.to_excel(file_path, index=False)
        apply_historical_formatting(file_path)
        
        # Close workbook explicitly
        wb = load_workbook(file_path)
        wb.close()
        
        return df

    except Exception as e:
        print(f"Error: {str(e)}")
        return df

if __name__ == "__main__":
    # For standalone use with fixed file
    EXCEL_PATH = r"ISM Manufacturing.xlsx"
    URL = "https://www.ismworld.org/supply-management-news-and-reports/reports/ism-report-on-business/pmi/february/"
    print("🚀 Starting ISM Report Update...")
    updated_df = update_excel_data(EXCEL_PATH, URL)
    print("\n✅ Process completed.")